import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
import os

class SistemaRestaurante:
    def __init__(self, archivo="pedidos.xlsx"):
        
        if not os.path.exists(archivo):
            self.crear_archivo_inicial(archivo)
        self.archivo = archivo
        self.workbook = openpyxl.load_workbook(archivo)
        self.sheet_pedidos = self.workbook["Pedidos"]
        self.sheet_menu = self.workbook["Menu"]

    def crear_archivo_inicial(self, archivo):
        
        wb = openpyxl.Workbook()
        
        
        sheet_pedidos = wb.active
        sheet_pedidos.title = "Pedidos"
        headers_pedidos = ["Mesa", "Fecha", "Hora Pedido", "Platos", "Total"]
        sheet_pedidos.append(headers_pedidos)
        
       
        sheet_menu = wb.create_sheet("Menu")
        headers_menu = ["Plato", "Precio"]
        sheet_menu.append(headers_menu)
        
        
        platos_ejemplo = [
            ("Pizza Margherita", 10.0),
            ("Hamburguesa", 8.5),
            ("Ensalada César", 6.0)
        ]
        for plato, precio in platos_ejemplo:
            sheet_menu.append([plato, precio])
        
        
        wb.save(archivo)

    def registrar_pedido(self):
        
        menu = {}
        for row in self.sheet_menu.iter_rows(min_row=2, values_only=True):
            plato, precio = row
            menu[plato] = precio
        
        
        print("Menú disponible:")
        for i, (plato, precio) in enumerate(menu.items(), 1):
            print(f"{i}. {plato} - ${precio}")
        
        
        mesa = input("Ingrese el número de mesa: ").strip()
        
        
        platos_seleccionados = []
        while True:
            seleccion = input("Ingrese el número del plato (0 para terminar): ")
            if seleccion == "0":
                break
            try:
                num = int(seleccion)
                if 1 <= num <= len(menu):
                    plato = list(menu.keys())[num - 1]
                    platos_seleccionados.append(plato)
                else:
                    print("Número inválido.")
            except ValueError:
                print("Por favor, ingrese un número.")
        
        if not platos_seleccionados:
            print("No se seleccionaron platos.")
            return
        
        
        total = sum(menu[plato] for plato in platos_seleccionados)
        
        
        fecha = datetime.now().strftime("%Y-%m-%d")
        hora_pedido = datetime.now().strftime("%H:%M:%S")
        platos_str = ", ".join(platos_seleccionados)
        self.sheet_pedidos.append([mesa, fecha, hora_pedido, platos_str, total])
        self.workbook.save(self.archivo)
        print(f"Pedido registrado para mesa {mesa}. Total: ${total:.2f}")

    